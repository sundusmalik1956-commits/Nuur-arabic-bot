# -*- coding: utf-8 -*-
"""
export_service.py
توليد ملف Excel بكل بيانات الطلاب المسجَّلين، لاستخدام إداري فقط (أمر /export
في bot.py، محصور على ADMIN_USER_IDS في config.py). لا علاقة له بمنطق تعليم
الطلاب — طبقة منفصلة تمامًا عن lesson_engine و database الأساسية، تقرأ منها فقط.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import database as db

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "exports")

COLUMNS = [
    ("user_id", "معرّف تيليجرام"),
    ("first_name", "الاسم"),
    ("username", "اسم المستخدم"),
    ("language", "اللغة"),
    ("level", "المستوى"),
    ("gender", "الجنس"),
    ("lesson_time", "وقت الدرس"),
    ("vacation_day_1", "يوم الإجازة الأول"),
    ("vacation_day_2", "يوم الإجازة الثاني"),
    ("current_lesson", "الدرس الحالي"),
    ("completed_lessons", "الدروس المكتملة"),
    ("subscription_status", "حالة الاشتراك"),
    ("active", "نشط"),
    ("created_at", "تاريخ التسجيل"),
]

WEEKDAY_ARABIC = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس",
                  4: "الجمعة", 5: "السبت", 6: "الأحد"}


def _format_value(key: str, value):
    if value is None:
        return ""
    if key in ("vacation_day_1", "vacation_day_2"):
        return WEEKDAY_ARABIC.get(value, value)
    if key == "active":
        return "نعم" if value else "لا"
    return value


def generate_students_excel() -> str:
    """يبني ملف Excel بكل بيانات الطلاب الحاليين، ويُعيد المسار الكامل للملف الناتج."""
    users = db.get_all_users()

    wb = Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name="Arial")

    for col_index, (_, header_label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, user in enumerate(users, start=2):
        for col_index, (key, _) in enumerate(COLUMNS, start=1):
            value = _format_value(key, user.get(key))
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center")

    for col_index in range(1, len(COLUMNS) + 1):
        column_letter = ws.cell(row=1, column=col_index).column_letter
        ws.column_dimensions[column_letter].width = 18

    ws.freeze_panes = "A2"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H-%M")
    file_path = os.path.join(OUTPUT_DIR, f"noor_bot_students_{timestamp}.xlsx")
    wb.save(file_path)

    return file_path
